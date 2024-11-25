import os
from flask import Flask, request, render_template, jsonify
import pytesseract
from PIL import Image
import pandas as pd
from openpyxl import Workbook

from fpdf import FPDF  # For PDF generation

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
EXCEL_FILE = "data.xlsx"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Ensure upload folder exists
try:                                                                                                                                                                                                                                                                                                                                                            
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
except OSError as error:
    print(f"Error creating directory: {error}")

# Ensure Excel file exists
if not os.path.exists(EXCEL_FILE):
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Details"])
        wb.save(EXCEL_FILE)
    except Exception as error:
        print(f"Error creating Excel file: {error}")

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
    file.save(filepath)

    # Perform OCR on the image
    try:
        text = pytesseract.image_to_string(Image.open(filepath))
    except Exception as error:
        return jsonify({"error": "Error performing OCR"}), 500

    # Extract and organize data
    lines = text.split("\n")
    organized_data = []
    for line in lines:
        if line.strip():
            parts = line.split(",")
            if len(parts) == 3:
                organized_data.append(parts)

    # Save data temporarily for preview
    return jsonify({"data": organized_data})

@app.route("/save", methods=["POST"])
def save_data():
    data = request.json.get("data", [])
    if not data:
        return jsonify({"error": "No data to save"}), 400

    try:
        df = pd.DataFrame(data, columns=["Name", "Date", "Details"])
        existing_data = pd.read_excel(EXCEL_FILE)
        merged_data = pd.concat([existing_data, df])
        merged_data.sort_values(by=["Name", "Date"], inplace=True)
        merged_data.to_excel(EXCEL_FILE, index=False)
    except Exception as error:
        return jsonify({"error": "Error saving data"}), 500

    return jsonify({"message": "Data saved successfully!"})

@app.route("/export_pdf", methods=["POST"])
def export_pdf():
    data = request.json.get("data", [])
    if not data:
        return jsonify({"error": "No data to export"}), 400

    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Add table headers
        pdf.cell(60, 10, "Name", 1)
        pdf.cell(40, 10, "Date", 1)
        pdf.cell(90, 10, "Details", 1)
        pdf.ln()

        # Add table rows
        for row in data:
            pdf.cell(60, 10, row[0], 1)
            pdf.cell(40, 10, row[1], 1)
            pdf.cell(90, 10, row[2], 1)
            pdf.ln()

        # Save PDF
        pdf_file = "extracted_data.pdf"
        pdf.output(pdf_file)
    except Exception as error:
        return jsonify({"error": "Error generating PDF"}), 500

    return jsonify({"message": "PDF generated successfully!", "file": pdf_file})

@app.route("/export_excel", methods=["POST"])
def export_excel():
    data = request.json.get("data", [])
    if not data:
        return jsonify({"error": "No data to export"}), 400

    try:
        df = pd.DataFrame(data, columns=["Name", "Date", "Details"])
        excel_file = "extracted_data.xlsx"
        df.to_excel(excel_file, index=False)
    except Exception as error:
        return jsonify({"error": "Error generating Excel file"}), 500

    return jsonify({"message": "Excel file generated successfully!", "file": excel_file})

@app.route("/delete_data", methods=["POST"])
def delete_data():
    try:
        os.remove(EXCEL_FILE)
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Details"])
        wb.save(EXCEL_FILE)
    except Exception as error:
        return jsonify({"error": "Error deleting data"}), 500

    return jsonify({"message": "Data deleted successfully!"})

if __name__ == "__main__":
    app.run(debug=True)